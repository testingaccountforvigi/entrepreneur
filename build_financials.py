# -*- coding: utf-8 -*-
"""Build a clean, 3-sheet, presentation-ready Aranya Carbon financial model."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- styles ----------
NAVY   = "0C2A1F"   # deep forest (titles)
GREEN  = "14532D"
EMER   = "12A06A"
GOLD   = "C2912E"
LIGHT  = "EAF2EC"   # section header bg
KEYBG  = "FCF3D6"   # key assumption highlight
WHITE  = "FFFFFF"
INKTX  = "1B2B23"
MUTED  = "5E7167"
BORDC  = "C9D6CC"

INR  = u'"\u20b9"#,##0'        # ₹#,##0
INR1 = u'"\u20b9"#,##0'        # whole rupees
USD  = u'"$"#,##0'
USD2 = u'"$"#,##0.00'
PCT  = '0%'
PCT1 = '0.0%'
NUM  = '#,##0'
NUM1 = '#,##0.0'

thin = Side(style="thin", color=BORDC)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def F(size=10, bold=False, color=INKTX, italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

def fill(c): return PatternFill("solid", fgColor=c)

def T(x):
    """Prevent text labels/notes that begin with '=' from being read as formulas."""
    if isinstance(x, str) and x.startswith("="):
        return "\u00a0" + x   # non-breaking space prefix -> treated as text, visually minimal
    return x

wb = Workbook()

# =====================================================================
# SHEET 1 — ASSUMPTIONS & SOURCES
# =====================================================================
s1 = wb.active
s1.title = "Assumptions & Sources"
s1.sheet_view.showGridLines = False
widths1 = {"A":2,"B":42,"C":16,"D":16,"E":58,"F":10}
for col,w in widths1.items(): s1.column_dimensions[col].width = w

A = {}  # name -> coordinate

def title(ws, row, text, sub=None):
    ws.cell(row=row, column=2, value=text).font = F(16, True, GREEN)
    if sub:
        ws.cell(row=row+1, column=2, value=sub).font = F(10, False, MUTED, italic=True)

def section(ws, row, text):
    for col in range(2, 7):
        cc = ws.cell(row=row, column=col)
        cc.fill = fill(LIGHT)
        cc.font = F(11, True, GREEN)
    ws.cell(row=row, column=2, value=text)

def headerrow(ws, row, cols):
    for i, h in enumerate(cols):
        cc = ws.cell(row=row, column=2+i, value=h)
        cc.font = F(10, True, WHITE)
        cc.fill = fill(GREEN)
        cc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cc.border = BORDER

def param(ws, row, label, value, unit="", note="", key=False, numfmt=None,
          is_formula=False, source=""):
    """Write a parameter row. Returns coordinate of the value cell (col C)."""
    lc = ws.cell(row=row, column=2, value=T(label)); lc.font = F(10, False, INKTX); lc.border = BORDER
    vc = ws.cell(row=row, column=3, value=value); vc.border = BORDER
    vc.alignment = Alignment(horizontal="right")
    if is_formula:
        vc.font = F(10, True, INKTX)              # formula = black bold
    else:
        vc.font = F(10, True, "1F4E79")           # input = blue
    if numfmt: vc.number_format = numfmt
    uc = ws.cell(row=row, column=4, value=unit); uc.font = F(10, False, MUTED); uc.border = BORDER
    nc = ws.cell(row=row, column=5, value=T(note)); nc.font = F(9, False, MUTED); nc.border = BORDER
    nc.alignment = Alignment(wrap_text=True, vertical="center")
    sc = ws.cell(row=row, column=6, value=source); sc.font = F(9, False, EMER); sc.border = BORDER
    if key:
        for col in range(2,7):
            ws.cell(row=row, column=col).fill = fill(KEYBG)
    return vc.coordinate

title(s1, 1, "ARANYA CARBON \u2014 FINANCIAL MODEL (2026)",
      "Community Forest Carbon Credit Platform \u2014 India  \u2022  Realistic, pre-seed / hackathon stage")
s1.cell(row=3, column=2,
        value="Legend:  BLUE = input you can change   \u2022   BLACK = formula   \u2022   YELLOW = key assumption to validate.  "
              "Change the exchange rate or price once here and the whole model updates.").font = F(9, False, MUTED, italic=True)

r = 5
section(s1, r, "A.  EXCHANGE RATE"); r += 1
headerrow(s1, r, ["Assumption", "Value", "Unit", "Notes", "Src"]); r += 1
A["fx"] = param(s1, r, "USD \u2192 INR exchange rate", 95, "INR per USD",
                "Latest mid-market rate, ~\u20b995.2 on 15-Jun-2026 (was \u20b983 in old model). All USD costs convert via this cell.",
                key=True, numfmt=INR, source="[S1]"); r += 1
r += 1

section(s1, r, "B.  CARBON CREDIT PRICING (per tCO\u2082e)"); r += 1
headerrow(s1, r, ["Assumption", "Value", "Unit", "Notes", "Src"]); r += 1
A["p_cons"] = param(s1, r, "Price \u2014 conservative", 8, "USD",
                    "Voluntary spot / generic REDD+ avg has been ~$6\u2013$9 in 2025\u201326.", numfmt=USD2, source="[S2]"); r += 1
A["p_real"] = param(s1, r, "Price \u2014 realistic (model base)", 13, "USD",
                    "High-integrity Indian community credit w/ co-benefits. Above ~$9 REDD+ spot, below the older $20\u2013$50 premiums that have since fallen.",
                    key=True, numfmt=USD2, source="[S2][S3]"); r += 1
A["p_prem"] = param(s1, r, "Price \u2014 premium / Article 6", 20, "USD",
                    "Co-benefit / Article 6 upside case only. Not used in base projections.", numfmt=USD2, source="[S3]"); r += 1
A["price_inr"] = param(s1, r, "Price \u2014 realistic (INR)", f"={A['p_real']}*{A['fx']}", "INR/tCO\u2082e",
                       "= realistic USD price \u00d7 exchange rate.", is_formula=True, numfmt=INR, source=""); r += 1
r += 1

section(s1, r, "C.  CARBON GENERATION (representative project)"); r += 1
headerrow(s1, r, ["Assumption", "Value", "Unit", "Notes", "Src"]); r += 1
A["ha"]   = param(s1, r, "Forest size \u2014 representative project", 200, "hectares",
                  "Mid-size community forest used for unit economics.", source=""); r += 1
A["seq"]  = param(s1, r, "Sequestration rate", 5, "tCO\u2082e/ha/yr",
                  "Conservative within IPCC tropical-forest range of 3\u20138; Indian dry-deciduous forests are at the lower end.",
                  key=True, source="[S4]"); r += 1
A["gross_cr"] = param(s1, r, "Gross credits / project / yr", f"={A['ha']}*{A['seq']}", "tCO\u2082e",
                      "= area \u00d7 sequestration rate.", is_formula=True, numfmt=NUM, source=""); r += 1
A["buffer"] = param(s1, r, "Permanence buffer (set-aside)", 0.15, "%",
                    "Credits held back as insurance against reversal (fire/loss). Verra norm 10\u201320%.",
                    numfmt=PCT, source="[S5]"); r += 1
A["saleable"] = param(s1, r, "Saleable credits / project / yr", f"={A['gross_cr']}*(1-{A['buffer']})", "tCO\u2082e",
                      "= gross credits \u00d7 (1 \u2212 buffer).", is_formula=True, numfmt=NUM, source=""); r += 1
r += 1

section(s1, r, "D.  VERIFICATION & REGISTRY COSTS"); r += 1
headerrow(s1, r, ["Assumption", "Value", "Unit", "Notes", "Src"]); r += 1
A["verra_reg"] = param(s1, r, "Verra registration fee (one-time)", 5000, "USD",
                       "Published Verra registration fee.", numfmt=USD, source="[S5]"); r += 1
A["valid"] = param(s1, r, "Validation audit (one-time, aggregated)", 12000, "USD",
                   "3rd-party validation, reduced from ~$25\u2013$40K via digital MRV + grouped projects.", numfmt=USD, source="[S6]"); r += 1
A["annual_verif"] = param(s1, r, "Annual verification \u2014 per project share", 700, "USD",
                          "One auditor visit shared across a cluster of ~20 grouped projects (the platform's core efficiency lever). Validate with first audits.",
                          key=True, numfmt=USD, source="[S6]"); r += 1
A["verra_ann"] = param(s1, r, "Verra annual maintenance", 500, "USD",
                       "Published Verra annual account fee.", numfmt=USD, source="[S5]"); r += 1
A["verif_passthru"] = param(s1, r, "Annual verification + registry pass-through (INR)",
                            f"=({A['annual_verif']}+{A['verra_ann']})*{A['fx']}", "INR/project/yr",
                            "= (annual verification + Verra fee) \u00d7 FX. Deducted from gross before community payout.",
                            is_formula=True, numfmt=INR, source=""); r += 1
A["onetime_aranya"] = param(s1, r, "One-time field onboarding cost (Aranya-borne)", 50000, "INR",
                            "Marginal cash cost per new project (field trip, data, consent). Team salaries are in OPEX.",
                            numfmt=INR, source=""); r += 1
A["industry_setup"] = param(s1, r, "Reference: industry one-time setup cost", 11300000, "INR",
                            "Solo origination for a small community forest historically ~\u20b91.1\u20131.2 Cr & 12\u201318 months (the barrier Aranya removes).",
                            numfmt=INR, source="[S6]"); r += 1
r += 1

section(s1, r, "E.  PLATFORM / REVENUE ASSUMPTIONS"); r += 1
headerrow(s1, r, ["Assumption", "Value", "Unit", "Notes", "Src"]); r += 1
A["comm"] = param(s1, r, "Aranya commission on credit sales", 0.18, "%",
                  "Platform origination fee. Industry range 15\u201325%; 18% mid-point.", key=True, numfmt=PCT, source="[S3]"); r += 1
A["monitor_fee"] = param(s1, r, "Monitoring subscription / project / yr", 100000, "INR",
                         "Annual fee for satellite MRV dashboard + audit-ready report.", numfmt=INR, source=""); r += 1
A["ngo_fee"] = param(s1, r, "NGO SaaS subscription", 12000, "INR/month",
                     "Per NGO partner for data app + monitoring + document tools.", numfmt=INR, source=""); r += 1
A["buyer_fee"] = param(s1, r, "Enterprise buyer subscription", 200000, "INR/year",
                       "Priority credit access + ESG reporting for corporate buyers.", numfmt=INR, source=""); r += 1
A["serv_cost"] = param(s1, r, "Aranya servicing cost / selling project / yr", 40000, "INR",
                       "Field servicing, support and per-project cloud/data allocation.", numfmt=INR, source=""); r += 1
A["mature_st"] = param(s1, r, "Mature-project sell-through", 0.90, "%",
                       "Share of saleable credits actually sold once a project is established.", numfmt=PCT, source=""); r += 1
r += 2

# Sources table
section(s1, r, "SOURCES & REFERENCES"); r += 1
headerrow(s1, r, ["Ref", "Source (organisation, year)", "", "URL", ""]); 
# merge note columns visually by just writing url in E
src_hdr = r; r += 1
sources = [
 ("[S1]", "USD/INR exchange rate \u2014 Wise / YCharts mid-market (15-Jun-2026): ~\u20b995.2 (1-yr ago ~\u20b985.6)",
  "https://wise.com/in/currency-converter/usd-to-inr-rate/history"),
 ("[S2]", "Carbon Credit Prices Today 2026 \u2014 voluntary spot avg ~$6; REDD+ ~$9/tonne",
  "https://carboncredits.com/carbon-prices-today/"),
 ("[S3]", "Abatable (2025) \u2014 new high-integrity REDD+ methodology floor price \u2265 $15/tonne",
  "https://abatable.com/blog/the-new-floor-price-for-redd-carbon-credits/"),
 ("[S4]", "IPCC 2006/2019 GHG Inventory Guidelines, Vol 4 (Forest Land) \u2014 tropical sequestration 3\u20138 tCO\u2082e/ha/yr",
  "https://www.ipcc-nggip.iges.or.jp/public/2006gl/vol4.html"),
 ("[S5]", "Verra \u2014 Verified Carbon Standard: registration ~$5,000; annual ~$500; permanence buffer 10\u201320%",
  "https://verra.org/programs/verified-carbon-standard/"),
 ("[S6]", "OPIS / Ecosystem Marketplace (2025) \u2014 REDD+ price benchmarks; origination & verification cost norms",
  "https://www.opis.com/blog/project-redd-struggles-amid-quality-push-future-of-voluntary-carbon-credits/"),
 ("[S7]", "Project documentation (details.txt) \u2014 business model, fee range 15\u201325%, community payout 70\u201382%",
  "Aranya Carbon project files"),
]
for ref, name, url in sources:
    s1.cell(row=r, column=2, value=ref).font = F(9, True, EMER)
    s1.cell(row=r, column=2).border = BORDER
    nc = s1.cell(row=r, column=3, value=name); nc.font = F(9, False, INKTX); nc.border = BORDER
    nc.alignment = Alignment(wrap_text=True, vertical="center")
    s1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    uc = s1.cell(row=r, column=5, value=url); uc.font = F(9, False, "1F4E79"); uc.border = BORDER
    uc.alignment = Alignment(wrap_text=True, vertical="center")
    if url.startswith("http"):
        uc.hyperlink = url
    s1.cell(row=r, column=6).border = BORDER
    r += 1
r += 1
s1.cell(row=r, column=2,
        value="Note: items marked yellow are the assumptions most worth validating with the first 3\u20135 pilot projects "
              "and live FX/credit-price quotes before fundraising.").font = F(9, False, MUTED, italic=True)

def sref(name):
    """Return a cross-sheet absolute reference to a Sheet-1 value cell."""
    coord = A[name]
    col = ''.join(ch for ch in coord if ch.isalpha())
    rownum = ''.join(ch for ch in coord if ch.isdigit())
    return f"'Assumptions & Sources'!${col}${rownum}"

# =====================================================================
# SHEET 2 — CAPEX & OPEX
# =====================================================================
s2 = wb.create_sheet("CAPEX & OPEX")
s2.sheet_view.showGridLines = False
for col,w in {"A":2,"B":46,"C":18,"D":18,"E":50}.items(): s2.column_dimensions[col].width = w

title(s2, 1, "CAPEX & OPEX \u2014 Aranya Carbon (Company Costs)",
      "What it costs to build and run the company. All figures in INR. Per-project external verification is shown in Sheet 3 (pass-through).")

r = 4
section(s2, r, "CAPEX \u2014 ONE-TIME STARTUP COSTS"); r += 1
headerrow(s2, r, ["Item", "Cost (INR)", "Cost (USD)", "Notes"]); r += 1
capex_start = r
capex = [
 ("Company incorporation, trademark & compliance", 75000, "One-time legal/registration setup."),
 ("Legal opinions \u2014 carbon rights / FRA / contracts", 300000, "Clarify community carbon ownership; buyer & panchayat agreements."),
 ("Platform / MVP build (beyond hackathon prototype)", 1000000, "Production web app, marketplace, MRV dashboard."),
 ("Field equipment (tablets, GPS, basic drone)", 250000, "Boundary mapping & inventory kit for field team."),
 ("Registry / developer account + first methodology", 200000, "Verra developer account setup & methodology onboarding."),
 ("Branding, website & launch marketing", 150000, "Identity, site, pilot outreach collateral."),
 ("Office setup & deposits", 200000, "Co-working deposit, furniture, basic infra."),
]
for item, cost, note in capex:
    s2.cell(row=r, column=2, value=item).font = F(10); s2.cell(row=r,column=2).border=BORDER
    vc = s2.cell(row=r, column=3, value=cost); vc.font = F(10, True, "1F4E79"); vc.number_format = INR; vc.border=BORDER
    uc = s2.cell(row=r, column=4, value=f"=C{r}/{sref('fx')}"); uc.font=F(10,False,INKTX); uc.number_format=USD; uc.border=BORDER
    nc = s2.cell(row=r, column=5, value=note); nc.font=F(9,False,MUTED); nc.border=BORDER; nc.alignment=Alignment(wrap_text=True,vertical="center")
    r += 1
capex_end = r-1
# contingency
s2.cell(row=r, column=2, value="Contingency (10%)").font=F(10); s2.cell(row=r,column=2).border=BORDER
vc=s2.cell(row=r,column=3,value=f"=SUM(C{capex_start}:C{capex_end})*0.1"); vc.font=F(10,True,INKTX); vc.number_format=INR; vc.border=BORDER
s2.cell(row=r,column=4,value=f"=C{r}/{sref('fx')}").number_format=USD; s2.cell(row=r,column=4).border=BORDER; s2.cell(row=r,column=4).font=F(10)
s2.cell(row=r,column=5,value="Standard buffer on a one-time budget.").font=F(9,False,MUTED); s2.cell(row=r,column=5).border=BORDER
cont_row=r; r+=1
# total capex
tc = s2.cell(row=r, column=2, value="TOTAL CAPEX (one-time)"); tc.font=F(11,True,WHITE); tc.fill=fill(GREEN); tc.border=BORDER
vc=s2.cell(row=r,column=3,value=f"=SUM(C{capex_start}:C{capex_end})+C{cont_row}"); vc.font=F(11,True,WHITE); vc.fill=fill(GREEN); vc.number_format=INR; vc.border=BORDER
uc=s2.cell(row=r,column=4,value=f"=C{r}/{sref('fx')}"); uc.font=F(11,True,WHITE); uc.fill=fill(GREEN); uc.number_format=USD; uc.border=BORDER
s2.cell(row=r,column=5).fill=fill(GREEN); s2.cell(row=r,column=5).border=BORDER
TOTAL_CAPEX = f"'CAPEX & OPEX'!$C${r}"
r += 2

section(s2, r, "OPEX \u2014 MONTHLY OPERATING COSTS (Year 1 baseline)"); r += 1
headerrow(s2, r, ["Item", "Monthly (INR)", "Annual (INR)", "Notes"]); r += 1
opex_start = r
opex = [
 ("Core team salaries (3 \u00d7 tech/ops, lean)", 135000, "Founders on minimal pre-seed stipends; lean build team."),
 ("Field officers (2 \u00d7 community onboarding)", 50000, "On-ground panchayat onboarding & data collection."),
 ("Cloud, data & APIs (hosting, GEE, Mapbox, AI)", 25000, "Vercel/Neon hosting, Earth Engine, maps, Gemini, payments."),
 ("Field operations (travel, fuel, per-diem)", 40000, "Site visits across pilot districts."),
 ("Sales & marketing", 20000, "Buyer outreach, NGO partnerships, content."),
 ("Office / co-working / utilities", 20000, "Workspace, internet, electricity."),
 ("Legal & accounting retainer", 15000, "Ongoing compliance, contracts, bookkeeping."),
 ("Software, tools & misc subscriptions", 10000, "Dev tools, SaaS, communications."),
 ("Contingency / buffer", 15000, "Unbudgeted month-to-month costs."),
]
for item, cost, note in opex:
    s2.cell(row=r, column=2, value=item).font=F(10); s2.cell(row=r,column=2).border=BORDER
    vc=s2.cell(row=r,column=3,value=cost); vc.font=F(10,True,"1F4E79"); vc.number_format=INR; vc.border=BORDER
    ac=s2.cell(row=r,column=4,value=f"=C{r}*12"); ac.font=F(10,False,INKTX); ac.number_format=INR; ac.border=BORDER
    nc=s2.cell(row=r,column=5,value=note); nc.font=F(9,False,MUTED); nc.border=BORDER; nc.alignment=Alignment(wrap_text=True,vertical="center")
    r += 1
opex_end = r-1
to=s2.cell(row=r,column=2,value="TOTAL OPEX"); to.font=F(11,True,WHITE); to.fill=fill(GREEN); to.border=BORDER
vc=s2.cell(row=r,column=3,value=f"=SUM(C{opex_start}:C{opex_end})"); vc.font=F(11,True,WHITE); vc.fill=fill(GREEN); vc.number_format=INR; vc.border=BORDER
ac=s2.cell(row=r,column=4,value=f"=SUM(D{opex_start}:D{opex_end})"); ac.font=F(11,True,WHITE); ac.fill=fill(GREEN); ac.number_format=INR; ac.border=BORDER
s2.cell(row=r,column=5,value="Monthly burn \u00d7 12 = Year-1 annual OPEX.").font=F(9,False,WHITE,italic=True); s2.cell(row=r,column=5).fill=fill(GREEN); s2.cell(row=r,column=5).border=BORDER
OPEX_MONTHLY=f"'CAPEX & OPEX'!$C${r}"
OPEX_ANNUAL_Y1=f"'CAPEX & OPEX'!$D${r}"
r += 2
s2.cell(row=r,column=2,value="Year-1 cash requirement (CAPEX + annual OPEX):").font=F(10,True,INKTX)
vc=s2.cell(row=r,column=3,value=f"={TOTAL_CAPEX}+{OPEX_ANNUAL_Y1}"); vc.font=F(11,True,GOLD); vc.number_format=INR
s2.cell(row=r,column=5,value="Indicative pre-seed ask (before revenue ramp). OPEX grows with headcount in Y2\u2013Y3 (see Sheet 3).").font=F(9,False,MUTED,italic=True)
s2.cell(row=r,column=5).alignment=Alignment(wrap_text=True)

# =====================================================================
# SHEET 3 — REVENUE MODEL & PROJECTIONS
# =====================================================================
s3 = wb.create_sheet("Revenue Model & Projections")
s3.sheet_view.showGridLines = False
for col,w in {"A":2,"B":44,"C":16,"D":16,"E":16,"F":40}.items(): s3.column_dimensions[col].width = w

title(s3, 1, "REVENUE MODEL & PROJECTIONS",
      "Realistic base case. Four revenue streams, per-project unit economics, and a conservative 3-year ramp.")

# ---- Revenue streams overview ----
r = 4
section(s3, r, "REVENUE STREAMS"); r += 1
headerrow(s3, r, ["Stream", "Basis", "Rate / Fee", "Notes"]); r += 1
streams = [
 ("1. Credit sale commission (primary)", "% of gross credit sales", f"={sref('comm')}", PCT, "Earned on every tonne sold via the marketplace."),
 ("2. Monitoring subscription", "Per project / year", f"={sref('monitor_fee')}", INR, "Satellite MRV dashboard + annual report."),
 ("3. NGO SaaS subscription", "Per NGO / month", f"={sref('ngo_fee')}", INR, "Platform access for partner NGOs."),
 ("4. Enterprise buyer subscription", "Per buyer / year", f"={sref('buyer_fee')}", INR, "Priority access + ESG reporting."),
]
for name, basis, rate, fmt, note in streams:
    s3.cell(row=r,column=2,value=name).font=F(10,True,INKTX); s3.cell(row=r,column=2).border=BORDER
    s3.cell(row=r,column=3,value=basis).font=F(9,False,MUTED); s3.cell(row=r,column=3).border=BORDER
    rc=s3.cell(row=r,column=4,value=rate); rc.font=F(10,True,INKTX); rc.number_format=fmt; rc.border=BORDER
    nc=s3.cell(row=r,column=5,value=note); nc.font=F(9,False,MUTED); nc.border=BORDER; nc.alignment=Alignment(wrap_text=True,vertical="center")
    r += 1
r += 1

# ---- Unit economics per project ----
section(s3, r, "UNIT ECONOMICS \u2014 PER PROJECT (mature, annual)"); r += 1
headerrow(s3, r, ["Line item", "Value", "% of gross", "Notes"]); r += 1
def ue(label, value, fmt=INR, pctof=None, note="", bold=False, isinput=False):
    global r
    lc=s3.cell(row=r,column=2,value=T(label)); lc.font=F(10, bold, INKTX); lc.border=BORDER
    vc=s3.cell(row=r,column=3,value=value); vc.number_format=fmt; vc.border=BORDER
    vc.font=F(10, bold, ("1F4E79" if isinput else INKTX)); vc.alignment=Alignment(horizontal="right")
    if pctof:
        pc=s3.cell(row=r,column=4,value=f"=C{r}/{pctof}"); pc.number_format=PCT1; pc.border=BORDER; pc.font=F(10,False,MUTED); pc.alignment=Alignment(horizontal="right")
    else:
        s3.cell(row=r,column=4).border=BORDER
    nc=s3.cell(row=r,column=5,value=T(note)); nc.font=F(9,False,MUTED); nc.border=BORDER; nc.alignment=Alignment(wrap_text=True,vertical="center")
    coord=f"C{r}"; r += 1; return coord

c_saleable = ue("Saleable credits / yr", f"={sref('saleable')}", NUM, note="From assumptions (after buffer).")
c_sold     = ue("Credits sold (mature sell-through)", f"={c_saleable}*{sref('mature_st')}", NUM, note="= saleable \u00d7 sell-through.")
c_price    = ue("Price per credit", f"={sref('price_inr')}", INR, note="Realistic price \u00d7 FX.")
c_gross    = ue("Gross credit sales", f"={c_sold}*{c_price}", INR, note="= credits sold \u00d7 price.", bold=True)
GROSS = f"$C${int(c_gross[1:])}"
c_comm     = ue("Aranya commission (18%)", f"={c_gross}*{sref('comm')}", INR, pctof=GROSS, note="Primary revenue.")
c_mon      = ue("Monitoring subscription", f"={sref('monitor_fee')}", INR, pctof=GROSS, note="Flat annual fee.")
c_arev     = ue("= Aranya revenue / project", f"={c_comm}+{c_mon}", INR, pctof=GROSS, note="Commission + monitoring.", bold=True)
c_verif    = ue("Less: verification + registry (pass-through)", f"=-{sref('verif_passthru')}", INR, pctof=GROSS, note="Paid to auditor/Verra; not Aranya margin.")
c_comm2    = ue("Less: Aranya commission", f"=-{c_comm}", INR, pctof=GROSS, note="Deducted from gross.")
c_payout   = ue("= Community payout", f"={c_gross}+{c_verif}+{c_comm2}", INR, pctof=GROSS, note="Residual to the panchayat (target 70\u201382% at scale).", bold=True)
c_serv     = ue("Aranya servicing cost / project", f"=-{sref('serv_cost')}", INR, pctof=GROSS, note="Field servicing + cloud allocation.")
c_contrib  = ue("= Aranya contribution / project / yr", f"={c_arev}+{c_serv}", INR, pctof=GROSS, note="Revenue minus direct servicing cost.", bold=True)
r += 1

# ---- 3-year projections ----
section(s3, r, "3-YEAR PROJECTION (REALISTIC RAMP)"); r += 1
headerrow(s3, r, ["Metric", "Year 1", "Year 2", "Year 3", "Notes / assumption"]); r += 1
hdr = r-1

def proj(label, y1, y2, y3, fmt=INR, isinput=False, bold=False, note="", fill_total=False):
    global r
    lc=s3.cell(row=r,column=2,value=T(label)); lc.border=BORDER
    cells=[]
    for i,val in enumerate([y1,y2,y3]):
        cc=s3.cell(row=r,column=3+i,value=val); cc.number_format=fmt; cc.border=BORDER
        cc.alignment=Alignment(horizontal="right")
        cc.font=F(10, bold, ("1F4E79" if isinput else INKTX))
        if fill_total:
            cc.fill=fill(LIGHT)
        cells.append(f"{get_column_letter(3+i)}{r}")
    lc.font=F(10, bold, INKTX)
    if fill_total: lc.fill=fill(LIGHT)
    nc=s3.cell(row=r,column=6,value=T(note)); nc.font=F(9,False,MUTED); nc.border=BORDER; nc.alignment=Alignment(wrap_text=True,vertical="center")
    rr=r; r += 1; return cells

# operating inputs (blue)
onboard = proj("Projects onboarded (cumulative)", 8, 20, 45, NUM, isinput=True,
               note="Conservative ramp; ~1 onboarding/month rising to ~2/month.")
selling = proj("Credit-generating projects (selling)", 2, 10, 32, NUM, isinput=True,
               note="Lags onboarding \u2014 registration/issuance takes ~9\u201318 months.")
sellthru= proj("Sell-through this year", 0.40, 0.75, 0.85, PCT, isinput=True,
               note="Share of saleable credits sold (ramps as buyer base matures).")
ngo     = proj("NGO SaaS partners", 2, 5, 14, NUM, isinput=True, note="Partner NGOs paying monthly SaaS.")
ngomon  = proj("NGO SaaS months billed", 6, 12, 12, NUM, isinput=True, note="Part-year in Y1.")
buyers  = proj("Enterprise buyer subscribers", 1, 3, 7, NUM, isinput=True, note="Corporates on paid subscription.")
opex_y  = proj("Operating expenses (OPEX)", f"={OPEX_ANNUAL_Y1}", 5500000, 8500000, INR, isinput=True,
               note="Y1 from Sheet 2; grows with headcount in Y2\u2013Y3.")
r += 1

# revenue (formulas)
rev_comm = proj("Credit commission revenue",
    f"={selling[0]}*{sref('saleable')}*{sellthru[0]}*{sref('price_inr')}*{sref('comm')}",
    f"={selling[1]}*{sref('saleable')}*{sellthru[1]}*{sref('price_inr')}*{sref('comm')}",
    f"={selling[2]}*{sref('saleable')}*{sellthru[2]}*{sref('price_inr')}*{sref('comm')}",
    INR, note="= selling \u00d7 saleable \u00d7 sell-through \u00d7 price \u00d7 18%.")
rev_mon = proj("Monitoring subscription revenue",
    f"={selling[0]}*{sref('monitor_fee')}", f"={selling[1]}*{sref('monitor_fee')}", f"={selling[2]}*{sref('monitor_fee')}",
    INR, note="= selling projects \u00d7 annual fee.")
rev_ngo = proj("NGO SaaS revenue",
    f"={ngo[0]}*{sref('ngo_fee')}*{ngomon[0]}", f"={ngo[1]}*{sref('ngo_fee')}*{ngomon[1]}", f"={ngo[2]}*{sref('ngo_fee')}*{ngomon[2]}",
    INR, note="= partners \u00d7 monthly fee \u00d7 months.")
rev_buy = proj("Enterprise buyer revenue",
    f"={buyers[0]}*{sref('buyer_fee')}", f"={buyers[1]}*{sref('buyer_fee')}", f"={buyers[2]}*{sref('buyer_fee')}",
    INR, note="= subscribers \u00d7 annual fee.")
tot_rev = proj("TOTAL REVENUE",
    f"=SUM({rev_comm[0]},{rev_mon[0]},{rev_ngo[0]},{rev_buy[0]})",
    f"=SUM({rev_comm[1]},{rev_mon[1]},{rev_ngo[1]},{rev_buy[1]})",
    f"=SUM({rev_comm[2]},{rev_mon[2]},{rev_ngo[2]},{rev_buy[2]})",
    INR, bold=True, fill_total=True, note="Sum of the four streams.")
r += 1

# costs (formulas)
cost_onb = proj("Onboarding cost (new projects)",
    f"={onboard[0]}*{sref('onetime_aranya')}",
    f"=({onboard[1]}-{onboard[0]})*{sref('onetime_aranya')}",
    f"=({onboard[2]}-{onboard[1]})*{sref('onetime_aranya')}",
    INR, note="= new projects this year \u00d7 onboarding cost.")
cost_serv = proj("Project servicing cost",
    f"={selling[0]}*{sref('serv_cost')}", f"={selling[1]}*{sref('serv_cost')}", f"={selling[2]}*{sref('serv_cost')}",
    INR, note="= selling projects \u00d7 servicing cost.")
cost_opex = proj("Operating expenses (OPEX)", f"={opex_y[0]}", f"={opex_y[1]}", f"={opex_y[2]}",
    INR, note="From operating inputs above.")
cost_capex = proj("CAPEX (one-time, Year 1)", f"={TOTAL_CAPEX}", 0, 0, INR, note="One-time company build (Sheet 2).")
tot_cost = proj("TOTAL COSTS",
    f"=SUM({cost_onb[0]},{cost_serv[0]},{cost_opex[0]},{cost_capex[0]})",
    f"=SUM({cost_onb[1]},{cost_serv[1]},{cost_opex[1]},{cost_capex[1]})",
    f"=SUM({cost_onb[2]},{cost_serv[2]},{cost_opex[2]},{cost_capex[2]})",
    INR, bold=True, fill_total=True, note="Onboarding + servicing + OPEX + CAPEX.")
r += 1

net = proj("NET PROFIT / (LOSS)",
    f"={tot_rev[0]}-{tot_cost[0]}", f"={tot_rev[1]}-{tot_cost[1]}", f"={tot_rev[2]}-{tot_cost[2]}",
    INR, bold=True, note="Realistic: invest in Y1\u2013Y2, approach breakeven by Y3.")
net_row = int(net[0][1:])
for i in range(3):
    s3.cell(row=net_row, column=3+i).font=F(11, True, GREEN)
    s3.cell(row=net_row, column=3+i).fill=fill(KEYBG)
s3.cell(row=net_row, column=2).fill=fill(KEYBG)

cum = proj("Cumulative net position",
    f"={net[0]}", f"={net[0]}+{net[1]}", f"={net[0]}+{net[1]}+{net[2]}",
    INR, note="Running total \u2014 indicates total capital needed before breakeven.")
r += 1
s3.cell(row=r, column=2,
        value="Takeaway: a deliberately conservative model \u2014 below-market credit price, conservative sequestration, "
              "slow issuance ramp. The platform invests through Years 1\u20132 and approaches breakeven in Year 3, "
              "consistent with a ~30-month path to Series A.").font=F(9,False,MUTED,italic=True)
s3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
s3.cell(row=r,column=2).alignment=Alignment(wrap_text=True, vertical="top")
s3.row_dimensions[r].height=42

# freeze panes for readability
s1.freeze_panes = "B5"
s3.freeze_panes = "B4"

wb.save("Aranya_Carbon_Financial_Model.xlsx")
print("Saved. Sheets:", wb.sheetnames)
